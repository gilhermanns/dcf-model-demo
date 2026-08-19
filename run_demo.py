import os
import click
import pandas as pd

from automated_dcf import DCFModel

@click.command()
@click.option('--tickers', default='AAPL,MSFT,TSLA', help='Comma-separated list of tickers')
def main(tickers):
    ticker_list = tickers.split(',')
    all_results = []
    output_dir = "outputs"
    charts_dir = os.path.join(output_dir, "charts")
    
    if not os.path.exists(charts_dir):
        os.makedirs(charts_dir)
        
    click.echo(f"Starting DCF Demo for: {', '.join(ticker_list)}")
    
    for ticker in ticker_list:
        try:
            click.echo(f"Processing {ticker}...")
            model = DCFModel(ticker)
            model.fetch_data()
            
            # Run DCF with 10-year projection
            results = model.run_dcf(years=10)
            all_results.append(results)
            
            # Export professional Excel
            excel_path = os.path.join(output_dir, f"{ticker}_DCF_Model.xlsx")
            model.export_to_excel(excel_path)
            
            # Generate charts
            model.plot_all_charts(output_dir=charts_dir)
            
            click.echo(f"Successfully completed {ticker}")
            
        except Exception as e:
            click.echo(f"Error processing {ticker}: {str(e)}", err=True)
            
    # Create Master Summary
    if all_results:
        click.echo("Creating Master Comparison...")
        summary_data = []
        for res in all_results:
            summary_data.append({
                'Ticker': res['ticker'],
                'Current Price': res['current_price'],
                'Implied Price': res['intrinsic_value'],
                'Upside (%)': res['upside'] * 100,
                'WACC (%)': res['wacc'] * 100,
                'Enterprise Value': res['enterprise_value'],
                'Equity Value': res['equity_value']
            })
        
        master_df = pd.DataFrame(summary_data)
        master_path = os.path.join(output_dir, "Master_Comparison.xlsx")
        
        from openpyxl.styles import Font, PatternFill
        with pd.ExcelWriter(master_path, engine='openpyxl') as writer:
            master_df.to_excel(writer, sheet_name='Comparison', index=False)
            ws = writer.sheets['Comparison']
            blue_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            white_font = Font(color='FFFFFF', bold=True)
            for cell in ws[1]:
                cell.fill = blue_fill
                cell.font = white_font
        
        click.echo(f"Master Comparison saved to {master_path}")
        
        # Print Markdown table for README
        click.echo("\n--- Summary Table ---")
        click.echo(master_df[['Ticker', 'Implied Price', 'Current Price', 'Upside (%)']].to_markdown(index=False))

if __name__ == "__main__":
    main()
